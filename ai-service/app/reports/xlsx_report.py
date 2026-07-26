"""
XLSX report via openpyxl. Deliberately structured as real spreadsheet data
(separate sheets, proper columns, conditional coloring) rather than a
table dump — an architect or reviewer is more likely to filter/sort this
than read it top to bottom, unlike the PDF/DOCX narrative versions.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SEVERITY_FILLS = {
    "CRITICAL": PatternFill(start_color="FFEF4444", end_color="FFEF4444", fill_type="solid"),
    "MAJOR": PatternFill(start_color="FFF59E0B", end_color="FFF59E0B", fill_type="solid"),
    "MINOR": PatternFill(start_color="FFEAB308", end_color="FFEAB308", fill_type="solid"),
}
HEADER_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)


def _autofit(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_xlsx(data: dict, output_path: str):
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    project = data["project"]
    ws["A1"] = "AI RERA Auditor — Compliance Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Project"; ws["B3"] = project.name
    ws["A4"] = "Location"; ws["B4"] = project.location or "—"
    ws["A5"] = "Compliance Score"; ws["B5"] = float(data["compliance_score"]) if data["compliance_score"] is not None else "Not yet analyzed"
    ws["A6"] = "Approval Probability"; ws["B6"] = float(data["approval_probability"]) if data["approval_probability"] is not None else "—"

    counts = {k: len(v) for k, v in data["violations_by_severity"].items()}
    ws["A8"] = "Critical"; ws["B8"] = counts["CRITICAL"]; ws["A8"].fill = SEVERITY_FILLS["CRITICAL"]
    ws["A9"] = "Major"; ws["B9"] = counts["MAJOR"]; ws["A9"].fill = SEVERITY_FILLS["MAJOR"]
    ws["A10"] = "Minor"; ws["B10"] = counts["MINOR"]; ws["A10"].fill = SEVERITY_FILLS["MINOR"]
    _autofit(ws, [22, 30])

    # --- Violations sheet ---
    ws2 = wb.create_sheet("Violations")
    headers = ["Severity", "Rule Code", "Category", "Description", "Detected Value", "Required Value", "Unit", "Floor", "Status"]
    for col, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    row = 2
    for v in data["violations"]:
        ws2.cell(row=row, column=1, value=v.severity).fill = SEVERITY_FILLS.get(v.severity)
        ws2.cell(row=row, column=2, value=v.rule.rule_code if v.rule else "—")
        ws2.cell(row=row, column=3, value=v.rule.category if v.rule else "—")
        ws2.cell(row=row, column=4, value=v.description)
        ws2.cell(row=row, column=5, value=float(v.detected_value) if v.detected_value is not None else None)
        ws2.cell(row=row, column=6, value=float(v.required_value) if v.required_value is not None else None)
        ws2.cell(row=row, column=7, value=v.rule.unit if v.rule else "")
        ws2.cell(row=row, column=8, value=v.floor.floor_number if v.floor else None)
        ws2.cell(row=row, column=9, value=v.status)
        row += 1
    ws2.freeze_panes = "A2"  # keep header visible when scrolling/filtering
    ws2.auto_filter.ref = f"A1:I{max(row - 1, 1)}"
    _autofit(ws2, [10, 26, 12, 50, 14, 14, 8, 8, 10])

    # --- Building data sheet ---
    ws3 = wb.create_sheet("Building Data")
    b_headers = ["Building", "Type", "Floors", "Height (m)", "Built-up Area (sqm)", "FAR", "Ground Coverage %"]
    for col, h in enumerate(b_headers, start=1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for i, b in enumerate(data["buildings"], start=2):
        ws3.cell(row=i, column=1, value=b.name)
        ws3.cell(row=i, column=2, value=b.building_type)
        ws3.cell(row=i, column=3, value=b.num_floors)
        ws3.cell(row=i, column=4, value=float(b.height_m) if b.height_m else None)
        ws3.cell(row=i, column=5, value=float(b.built_up_area_sqm) if b.built_up_area_sqm else None)
        ws3.cell(row=i, column=6, value=float(b.far_calculated) if b.far_calculated else None)
        ws3.cell(row=i, column=7, value=float(b.ground_coverage_pct) if b.ground_coverage_pct else None)
    _autofit(ws3, [20, 16, 10, 12, 20, 10, 18])

    # --- AI Suggestions sheet ---
    if data["suggestions"]:
        ws4 = wb.create_sheet("AI Suggestions")
        ws4["A1"] = "Category"; ws4["B1"] = "Suggestion"
        ws4["A1"].fill = HEADER_FILL; ws4["A1"].font = HEADER_FONT
        ws4["B1"].fill = HEADER_FILL; ws4["B1"].font = HEADER_FONT
        for i, s in enumerate(data["suggestions"], start=2):
            ws4.cell(row=i, column=1, value=s.category or "GENERAL")
            cell = ws4.cell(row=i, column=2, value=s.suggestion_text)
            cell.alignment = Alignment(wrap_text=True)
        _autofit(ws4, [16, 80])

    wb.save(output_path)
